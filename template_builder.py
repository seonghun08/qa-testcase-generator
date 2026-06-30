"""
QA Test Case 양식 빌더 v2
- 원본 색상 정확히 매칭 (연녹색 헤더 #E2EEDA, 노란 띠 #FFD966)
- 서버환경/수행환경 행 제거
- 모든 셀 세로 가운데 정렬
- ITS 컬럼 제거

사용법:
  - 모듈로 import: from template_builder import build_workbook
  - 직접 실행: python template_builder.py  (샘플 xlsx 생성)
"""
import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === 색상 (원본 매칭) ===
HEADER_BG = 'E2EEDA'   # 연녹색 (theme:9 tint:0.8)
ACCENT_BG = 'FFD966'   # 노란 띠 (theme:7 tint:0.4)
BORDER_COLOR = '8EA9DB'  # 연한 회색 (원본 thin border)

# === 폰트 ===
KF = '맑은 고딕'

TITLE_FONT = Font(name=KF, size=18, bold=True, color='000000')
HEAD_FONT = Font(name=KF, size=9, bold=True, color='000000')
NORMAL_FONT = Font(name=KF, size=9, color='000000')
DATA_FONT = Font(name=KF, size=9, color='000000')

# === 상태값별 스타일 (UI/로그/DB 영역 색상) ===
STATUS_STYLES = {
    'Pass':    {'bg': 'C6EFCE', 'font_color': '006100'},
    'Fail':    {'bg': 'FFC7CE', 'font_color': '9C0006'},
    'Fixed':   {'bg': 'FFEB9C', 'font_color': '9C5700'},
    'Blocked': {'bg': 'E4DFEC', 'font_color': '5B3970'},
    'N/A':     {'bg': 'D9D9D9', 'font_color': '595959'},
    'NotTest': {'bg': 'FFFFFF', 'font_color': '000000'},
}

# === 채우기 ===
HEADER_FILL = PatternFill('solid', start_color=HEADER_BG, end_color=HEADER_BG)
ACCENT_FILL = PatternFill('solid', start_color=ACCENT_BG, end_color=ACCENT_BG)

# === 보더 ===
thin = Side(style='thin', color='808080')
medium = Side(style='medium', color='606060')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_HEADER_LEFT = Border(left=medium, right=thin, top=thin, bottom=thin)
BORDER_TITLE = Border(top=medium, left=medium, right=medium, bottom=None)
NO_BORDER = Border()

# === 정렬 (모든 셀 세로 가운데) ===
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

# === 로그/DB 접속 정보 (사내 환경값은 외부 config로 분리) ===
# 실제 서버 IP/경로/SID 등은 git에 올리지 않는 log_db_config.json 에 둔다.
# (없으면 log_db_config.example.json → 내장 기본 플레이스홀더 순으로 대체)
_CFG_DIR = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_LOGDB = {
    "LOG_APP_LABEL": "WAS 접속",
    "LOG_SERVER": "<WAS_HOST:PORT>",
    "LOG_DIR": "<TOMCAT_LOG_DIR>",
    "DB_LABEL": "테스트DB 접속",
    "DB_SERVER": "<DB_HOST:PORT>",
    "DB_SID": "<DB_SID>",
}


def _load_logdb_config():
    for fn in ("log_db_config.json", "log_db_config.example.json"):
        path = os.path.join(_CFG_DIR, fn)
        if os.path.isfile(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                merged = dict(_DEFAULT_LOGDB)
                merged.update({k: v for k, v in cfg.items() if k in _DEFAULT_LOGDB})
                return merged
            except (OSError, json.JSONDecodeError):
                pass
    return dict(_DEFAULT_LOGDB)


_LOGDB = _load_logdb_config()


def _apply_logdb(template):
    """@KEY@ 마커를 config 값으로 치환한다(개발자 입력용 {..} 플레이스홀더는 보존)."""
    for key, val in _LOGDB.items():
        template = template.replace(f"@{key}@", str(val))
    return template


# === 로그/DB 템플릿 (@..@ 는 config 치환, {..} 는 개발자가 채우는 자리) ===
LOG_TEMPLATE = _apply_logdb(
    "- @LOG_APP_LABEL@ (개인계정 사용)\n"
    "  ▷ @LOG_SERVER@\n"
    "- 디렉토리 이동\n"
    "  ▷ @LOG_DIR@\n"
    "- 로그 검색\n"
    "  ▷ grep -i \"{mapper_id}\" ./catalina.out\n"
    "     > 출력된 로그내 쿼리 확인\n"
    "  또는\n"
    "  ▷ tail -f catalina.out\n"
    "     > tail 상태에서 화면에서 조회 버튼 클릭\n"
    "     > 출력된 로그내 쿼리 확인"
)

DB_TEMPLATE_SELECT = _apply_logdb(
    "- @DB_LABEL@ (개인계정 사용)\n"
    "  ▷ @DB_SERVER@ (SID:@DB_SID@)\n"
    "- 로그에서 출력된 쿼리 복사 후 DB툴에서 실행 시 정상 확인 및 데이터 비교\n\n"
    "  [쿼리 예시]\n"
    "  SELECT * FROM {스키마}.{테이블명}\n"
    "  WHERE {조건컬럼} = '{예시값}'\n\n"
    "  [컬럼 설명]\n"
    "  {컬럼명}: {설명}"
)

DB_TEMPLATE_INSERT = _apply_logdb(
    "- @DB_LABEL@ (개인계정 사용)\n"
    "  ▷ @DB_SERVER@ (SID:@DB_SID@)\n"
    "- 조회 쿼리 실행하여 저장한 데이터 확인\n\n"
    "  [쿼리 예시]\n"
    "  SELECT * FROM {스키마}.{테이블명}\n"
    "  WHERE {PK컬럼} = '{자동채번값}'\n\n"
    "  [컬럼 설명]\n"
    "  {PK컬럼}: {설명}"
)

DB_TEMPLATE_UPDATE = _apply_logdb(
    "- @DB_LABEL@ (개인계정 사용)\n"
    "  ▷ @DB_SERVER@ (SID:@DB_SID@)\n"
    "- 조회 쿼리 실행하여 수정한 데이터 확인\n\n"
    "  [쿼리 예시]\n"
    "  SELECT {수정컬럼}, {상태컬럼} FROM {스키마}.{테이블명}\n"
    "  WHERE {PK컬럼} = '{식별값}'\n\n"
    "  [컬럼 설명]\n"
    "  {수정컬럼}: {설명}\n"
    "  {상태컬럼}: {상태값}"
)

DB_TEMPLATES = {
    'select': DB_TEMPLATE_SELECT,
    'insert': DB_TEMPLATE_INSERT,
    'update': DB_TEMPLATE_UPDATE,
}

# 기본 DB 안내 (미지정 시 모든 케이스 공통 적용)
DB_TEMPLATE_DEFAULT = _apply_logdb(
    "- @DB_LABEL@ (개인계정 사용)\n"
    "  ▷ @DB_SERVER@ (SID:@DB_SID@)\n"
    "- 로그에서 출력된 쿼리 복사 후 DB툴에서 실행 시 정상 확인 및 데이터 비교"
)

# === 시스템 타입 매핑 ===
SYSTEM_TYPES = {
    'BO': 'BO\n(내부관리자)',
    'SHOP': 'SHOP\n(가맹점관리자)',
    'OSS': 'OSS\n(운영지원시스템)',
}

# 시스템 타입별 한글 명칭 (압축 라벨 생성용)
SYSTEM_NAMES = {
    'BO': '내부관리자',
    'SHOP': '가맹점관리자',
    'OSS': '운영지원시스템',
}


def normalize_system_types(value, default_types):
    """시스템 타입 입력을 유효한 키 리스트로 정규화한다.

    문자열('BO', 'BO/SHOP', 'BO,SHOP')·리스트(['BO','SHOP']) 모두 허용.
    유효한 값이 없으면 default_types를 반환한다.
    """
    if not value:
        return list(default_types)

    if isinstance(value, str):
        parts = re.split(r'[\/,\s]+', value)
    elif isinstance(value, (list, tuple)):
        parts = []
        for v in value:
            parts.extend(re.split(r'[\/,\s]+', str(v)))
    else:
        return list(default_types)

    seen, out = set(), []
    for p in parts:
        key = p.strip().upper()
        if key in SYSTEM_NAMES and key not in seen:
            seen.add(key)
            out.append(key)
    return out or list(default_types)


def _common_suffix(names):
    """문자열 리스트의 공통 접미사를 반환한다 (없으면 '')."""
    if len(names) < 2:
        return ''
    suffix = ''
    for chars in zip(*[n[::-1] for n in names]):
        if len(set(chars)) == 1:
            suffix = chars[0] + suffix
        else:
            break
    return suffix


def make_system_label(types):
    """시스템 타입 키 리스트를 '테스트 대상' 셀용 압축 라벨로 만든다.

    예) ['BO']        -> 'BO\\n(내부관리자)'
        ['BO','SHOP'] -> 'BO/SHOP\\n(내부/가맹점관리자)'   (공통 접미사 '관리자' 병합)
        ['BO','OSS']  -> 'BO/OSS\\n(내부관리자/운영지원시스템)'
    """
    if not types:
        types = ['BO']
    codes = '/'.join(types)
    names = [SYSTEM_NAMES[t] for t in types]

    if len(names) == 1:
        label = names[0]
    else:
        suffix = _common_suffix(names)
        if suffix:
            stems = [n[:-len(suffix)] for n in names]
            label = '/'.join(stems[:-1] + [stems[-1] + suffix])
        else:
            label = '/'.join(names)

    return f"{codes}\n({label})"

# === 헬퍼 함수 ===

def status_font(status):
    fc = STATUS_STYLES.get(status, {}).get('font_color', '000000')
    return Font(name=KF, size=9, bold=True, color=fc)

def status_fill(status):
    bg = STATUS_STYLES.get(status, {}).get('bg')
    return PatternFill('solid', start_color=bg, end_color=bg) if bg else None

def write_status(ws, coord, status):
    c = ws[coord]
    c.value = status
    c.font = status_font(status) if status else DATA_FONT
    fill = status_fill(status) if status else None
    if fill:
        c.fill = fill
    c.alignment = CENTER
    c.border = BORDER
    return c

def style_range_cells(ws, cell_range, font=None, fill=None, alignment=None, border=None):
    if ':' in cell_range:
        rows = ws[cell_range]
        if not isinstance(rows, tuple):
            rows = (rows,)
        for row in rows:
            if not isinstance(row, tuple):
                row = (row,)
            for cell in row:
                if font: cell.font = font
                if fill: cell.fill = fill
                if alignment: cell.alignment = alignment
                if border: cell.border = border
    else:
        cell = ws[cell_range]
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border

def write(ws, coord, value, font=DATA_FONT, fill=None, alignment=CENTER, border=BORDER):
    c = ws[coord]
    c.value = value
    c.font = font
    if fill: c.fill = fill
    c.alignment = alignment
    if border: c.border = border
    return c


# ============================================================
# 메인 빌더 함수
# ============================================================

def build_workbook(jira_num, title_base, test_cases, system_type='BO',
                   developer_name='', tester_name='', qa_name=''):
    """
    QA 테스트 케이스 xlsx 워크북을 생성한다.

    Args:
        jira_num: Jira 티켓 번호 (예: "PGBO-1001")
        title_base: 문서 제목 (예: "가맹점 정산 내역 관리 화면 개발 요청")
        test_cases: 테스트 케이스 리스트. 각 항목은 dict:
            {
                'id': 'TB_01_001',
                'precond': '사전조건 텍스트',
                'step': '조회|등록|수정|삭제',
                'category': '구분',
                'item': '항목',
                'procedure': '수행절차',
                'expected': '기대결과',
                'log': '로그 (선택, 없으면 기본 템플릿)',
                'db': 'DB (선택, 없으면 log_db_type 기반 템플릿)',
                'log_db_type': 'select|insert|update (db 미지정 시 사용)',
            }
        system_type: 시스템 타입 키 ('BO', 'SHOP', 'OSS')
        developer_name: 개발자명 (헤더에 표시)
        tester_name: 테스터명 (헤더에 표시)
        qa_name: QA명 (헤더에 표시)

    Returns:
        (wb, file_name): openpyxl Workbook 객체와 파일명
    """
    title_txt = f"TB_TC_{jira_num} {title_base}"
    # 파일명에 쓸 수 없는 문자(\ / : * ? " < > |)는 _로 치환하고 공백도 _로,
    # 연속된 _는 하나로 합친 뒤 양끝 _를 제거한다. (제목의 '/' 등이 경로 구분자로
    # 오인돼 저장이 실패하는 문제 방지. B2 내부 제목 title_txt는 원본 그대로 유지)
    safe_title = re.sub(r'[\\/:*?"<>|\s]+', '_', title_base).strip('_')
    file_name = f"TB_TC_{jira_num}_{safe_title}.xlsx"
    # 문서 기본 시스템 타입 (케이스가 별도 지정 안 하면 이 값으로 fallback)
    default_types = normalize_system_types(system_type, ['BO'])

    wb = Workbook()

    # === Sheet 1: 문서변경이력 ===
    ws1 = wb.active
    ws1.title = "문서변경이력"

    write(ws1, 'B2', '문서변경 이력', font=TITLE_FONT, alignment=CENTER, border=NO_BORDER)
    ws1.merge_cells('B2:K2')
    style_range_cells(ws1, 'B2:K2', font=TITLE_FONT, fill=HEADER_FILL, alignment=CENTER)
    ws1.row_dimensions[2].height = 36

    ws1.merge_cells('B3:K3')
    style_range_cells(ws1, 'B3:K3', fill=ACCENT_FILL, border=NO_BORDER)
    ws1.row_dimensions[3].height = 8

    ws1.merge_cells('C6:I6')
    write(ws1, 'B6', '#', font=HEAD_FONT, fill=HEADER_FILL)
    style_range_cells(ws1, 'C6:I6', font=HEAD_FONT, fill=HEADER_FILL, alignment=CENTER, border=BORDER)
    ws1['C6'].value = '변경내용'
    write(ws1, 'J6', '작성자', font=HEAD_FONT, fill=HEADER_FILL)
    write(ws1, 'K6', '작성일', font=HEAD_FONT, fill=HEADER_FILL)
    ws1.row_dimensions[6].height = 22

    ws1.merge_cells('C7:I7')
    write(ws1, 'B7', 1, font=DATA_FONT, alignment=CENTER)
    style_range_cells(ws1, 'C7:I7', font=DATA_FONT, alignment=LEFT_CENTER, border=BORDER)
    ws1['C7'].value = '최초 작성'
    write(ws1, 'J7', '', font=DATA_FONT, alignment=CENTER)
    write(ws1, 'K7', '', font=DATA_FONT, alignment=CENTER)
    ws1.row_dimensions[7].height = 22

    ws1.merge_cells('C8:I8')
    write(ws1, 'B8', None, font=DATA_FONT, alignment=CENTER)
    style_range_cells(ws1, 'C8:I8', font=DATA_FONT, alignment=LEFT_CENTER, border=BORDER)
    write(ws1, 'J8', None, font=DATA_FONT, alignment=CENTER)
    write(ws1, 'K8', None, font=DATA_FONT, alignment=CENTER)
    ws1.row_dimensions[8].height = 22

    widths1 = {'A': 2, 'B': 5, 'C': 18, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 8, 'I': 8, 'J': 15, 'K': 15}
    for col, w in widths1.items():
        ws1.column_dimensions[col].width = w

    # === Sheet 2: TestCase ===
    ws2 = wb.create_sheet("TestCase")

    col_widths = {
        'A': 2, 'B': 14, 'C': 7, 'D': 13, 'E': 26, 'F': 8, 'G': 16, 'H': 18,
        'I': 38, 'J': 35, 'K': 48, 'L': 45,
        'M': 8, 'N': 8, 'O': 8, 'P': 8, 'Q': 8, 'R': 8, 'S': 8, 'T': 8, 'U': 8,
        'V': 14, 'W': 22,
    }
    for col, w in col_widths.items():
        ws2.column_dimensions[col].width = w

    # 제목 (Row 2)
    ws2.merge_cells('B2:W2')
    title_align = Alignment(horizontal='left', vertical='center', indent=2)
    write(ws2, 'B2', title_txt, font=TITLE_FONT, fill=HEADER_FILL, alignment=title_align, border=NO_BORDER)
    style_range_cells(ws2, 'B2:W2', font=TITLE_FONT, fill=HEADER_FILL, alignment=title_align)
    ws2.row_dimensions[2].height = 40

    # 노란 띠 (Row 3)
    ws2.merge_cells('B3:W3')
    style_range_cells(ws2, 'B3:W3', fill=ACCENT_FILL, border=NO_BORDER)
    ws2.row_dimensions[3].height = 8

    # Row 4: 빈 행
    ws2.row_dimensions[4].height = 12

    # 헤더 행 (Row 5-6)
    ws2.row_dimensions[5].height = 28
    ws2.row_dimensions[6].height = 22

    single_headers = [
        ('B', 'TestCaseID'), ('C', '우선순위'), ('D', '테스트 대상'), ('E', '사전조건'),
        ('I', '수행절차'), ('J', '기대결과'), ('K', '로그'), ('L', 'DB'),
        ('V', 'JIRA'), ('W', '비고'),
    ]
    for col, text in single_headers:
        rng = f'{col}5:{col}6'
        ws2.merge_cells(rng)
        style_range_cells(ws2, rng, font=HEAD_FONT, fill=HEADER_FILL, alignment=CENTER, border=BORDER)
        ws2[f'{col}5'].value = text

    # 테스트 항목 그룹 (F-H)
    ws2.merge_cells('F5:H5')
    style_range_cells(ws2, 'F5:H5', font=HEAD_FONT, fill=HEADER_FILL, alignment=CENTER, border=BORDER)
    ws2['F5'].value = '테스트 항목'
    for col, text in [('F', '단계'), ('G', '구분'), ('H', '항목')]:
        write(ws2, f'{col}6', text, font=HEAD_FONT, fill=HEADER_FILL, alignment=CENTER)

    # 평가자 그룹 (M-U)
    dev_label = f'개발자 ({developer_name})' if developer_name else '개발자 ()'
    tst_label = f'테스터 ({tester_name})' if tester_name else '테스터 ()'
    qa_label = f'QA ({qa_name})' if qa_name else 'QA ()'
    groups = [('M', 'O', dev_label), ('P', 'R', tst_label), ('S', 'U', qa_label)]
    for start, end, name in groups:
        rng = f'{start}5:{end}5'
        ws2.merge_cells(rng)
        style_range_cells(ws2, rng, font=HEAD_FONT, fill=HEADER_FILL, alignment=CENTER, border=BORDER)
        ws2[f'{start}5'].value = name
        start_idx = ord(start) - ord('A') + 1
        for offset, sub in enumerate(['UI', '로그', 'DB']):
            cl = get_column_letter(start_idx + offset)
            write(ws2, f'{cl}6', sub, font=HEAD_FONT, fill=HEADER_FILL, alignment=CENTER)

    # Row 7: 스페이서
    ws2.row_dimensions[7].height = 6

    # 테스트 케이스 데이터 (Row 8+)
    TC_ROW_HEIGHT = 230
    TC_START_ROW = 8

    for idx, tc in enumerate(test_cases):
        r = TC_START_ROW + idx
        ws2.row_dimensions[r].height = TC_ROW_HEIGHT

        # 로그/DB는 미지정 시 기본 안내 템플릿으로 채운다.
        log_val = tc.get('log') or LOG_TEMPLATE
        db_val = tc.get('db') or DB_TEMPLATE_DEFAULT

        # 케이스별 시스템 타입 (BO/SHOP 동시 적용 등). 없으면 문서 기본값.
        row_types = normalize_system_types(tc.get('system_type'), default_types)
        write(ws2, f'B{r}', tc['id'], alignment=CENTER)
        write(ws2, f'C{r}', 'P0', alignment=CENTER)
        write(ws2, f'D{r}', make_system_label(row_types), alignment=CENTER)
        write(ws2, f'E{r}', tc['precond'], alignment=LEFT_CENTER)
        write(ws2, f'F{r}', tc['step'], alignment=CENTER)
        write(ws2, f'G{r}', tc['category'], alignment=CENTER)
        write(ws2, f'H{r}', tc['item'], alignment=CENTER)
        write(ws2, f'I{r}', tc['procedure'], alignment=LEFT_CENTER)
        write(ws2, f'J{r}', tc['expected'], alignment=LEFT_CENTER)
        write(ws2, f'K{r}', log_val, alignment=LEFT_CENTER)
        write(ws2, f'L{r}', db_val, alignment=LEFT_CENTER)

        # 결과(개발자/테스터/QA UI·로그·DB)는 테스트 전이므로 전부 공란으로 둔다.
        for col in ('M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U'):
            write_status(ws2, f'{col}{r}', None)
        # JIRA, 비고: 빈 칸
        write(ws2, f'V{r}', None, alignment=CENTER)
        write(ws2, f'W{r}', None, alignment=LEFT_CENTER)

    ws2.freeze_panes = 'C8'

    # === Sheet 3: 참고_TestCase 결과 ===
    ws3 = wb.create_sheet("참고_TestCase 결과")

    ws3.merge_cells('B2:G2')
    style_range_cells(ws3, 'B2:G2', font=TITLE_FONT, fill=HEADER_FILL, alignment=CENTER)
    ws3['B2'].value = 'Test Case 결과 코드 정의'
    ws3.row_dimensions[2].height = 36

    ws3.merge_cells('B3:G3')
    style_range_cells(ws3, 'B3:G3', fill=ACCENT_FILL, border=NO_BORDER)
    ws3.row_dimensions[3].height = 8

    ref_headers = [('B', '결과'), ('C', '결과 흐름도'), ('D', '사유'), ('E', '커버리지'), ('F', 'QA 담당자'), ('G', '테스터')]
    for col, text in ref_headers:
        write(ws3, f'{col}5', text, font=HEAD_FONT, fill=HEADER_FILL, alignment=CENTER)
    ws3.row_dimensions[5].height = 24

    ref_data = [
        ('Pass', '', '기대결과 일치', '포함', '-', '선택: 비고 작성'),
        ('Fixed', 'Fail > Fixed', '1. Fail > Fixed: 이슈 수정\n2. Blocked > Fixed: 확인 불가 > 정상 동작 확인', '포함', '-', '필수: 비고, JIRA 작성'),
        ('Fail', '', '기대결과 불일치', '포함', '-', '-'),
        ('Blocked', 'Blocked > Fixed', '1. 현재 확인 불가 (상태)\n2. 상위 TestCase Fail로 인해 확인 불가\n3. 미구현 (선공유)', '미포함', '필수: 1, 3에 대한 비고 작성', '필수: 2에 대한 비고 작성'),
        ('N/A', '', '테스트 범위 제외', '미포함', '필수: 비고 작성', '-'),
    ]
    for i, row in enumerate(ref_data):
        r = 6 + i
        ws3.row_dimensions[r].height = 55
        for col_idx, val in enumerate(row):
            col = chr(ord('B') + col_idx)
            align = CENTER if col_idx in [0, 1, 3] else LEFT_CENTER
            if col_idx == 0 and val in STATUS_STYLES:
                write_status(ws3, f'{col}{r}', val)
            else:
                write(ws3, f'{col}{r}', val, alignment=align)

    widths3 = {'A': 2, 'B': 10, 'C': 16, 'D': 38, 'E': 10, 'F': 24, 'G': 24}
    for col, w in widths3.items():
        ws3.column_dimensions[col].width = w

    return wb, file_name


def save_workbook(wb, file_name, out_dir):
    """워크북을 지정 디렉토리에 저장한다."""
    import os
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, file_name)
    wb.save(out_path)
    return out_path


# ============================================================
# 직접 실행 시 샘플 생성
# ============================================================
if __name__ == '__main__':
    import os

    PRECOND_LIST = "내부직원 계정 로그인\n\nPG 정산 관리 > 가맹점 정산 내역"
    PRECOND_DETAIL = "내부직원 계정 로그인\n\nPG 정산 관리 > 가맹점 정산 내역 > 상세"

    sample_cases = [
        {'id': 'TB_01_001', 'precond': PRECOND_LIST, 'step': '조회', 'category': '정산 내역 목록 조회',
         'item': '검색 조건 별\n목록 조회',
         'procedure': "1. 검색 조건 입력\n 1-1. 가맹점명, 가맹점번호, 정산일자 범위, 정산상태 등\n2. 조회 버튼 클릭\n3. 검색 조건 초기화 버튼 동작 확인",
         'expected': "1. 검색 조건에 맞는 정산 내역 정상 조회\n\n2. 정산 상태별 색상 구분\n - 대기(W): 검정\n - 완료(C): 파랑\n - 보류(H): 빨강\n - 취소(X): 회색\n\n3. 초기 진입 시 당일 기준 7일 데이터 자동 조회",
         'log_db_type': 'select'},
        {'id': 'TB_01_002', 'precond': PRECOND_LIST, 'step': '조회', 'category': '정산 내역 목록 조회',
         'item': '정렬 및 페이징',
         'procedure': "1. 컬럼 헤더 클릭 시 오름차순/내림차순 정렬 확인\n 1-1. 정산일자, 정산금액, 가맹점명 등 주요 컬럼\n2. 페이지당 노출 건수 변경 (10/30/50/100건)\n3. 페이지 이동 (이전, 다음, 직접 이동) 동작 확인",
         'expected': "1. 정렬 시 화면 및 데이터 정상 변경\n\n2. 페이지당 건수 변경 시 1페이지부터 재조회\n\n3. 총 건수 / 현재 페이지 표시 정상",
         'log_db_type': 'select'},
        {'id': 'TB_01_003', 'precond': PRECOND_LIST, 'step': '수정', 'category': '정산 보류 처리',
         'item': '정산 보류 처리',
         'procedure': "1. 정산 상태가 '대기(W)'인 건 체크박스 선택\n 1-1. 다중 선택 가능\n2. 하단 '정산 보류' 버튼 클릭\n3. 보류 사유 입력 팝업 노출\n 3-1. 사유 미입력 시 저장 불가 (필수)\n4. 사유 입력 후 확인 클릭",
         'expected': "1. '대기(W)' 상태가 아닌 건 선택 시 경고 메시지\n\n2. 보류 처리 후 정산 상태 '보류(H)' 변경\n\n3. 목록 재조회 시 빨간색 표시\n\n4. 보류 사유 및 처리자 정보 기록",
         'log_db_type': 'update'},
        {'id': 'TB_01_004', 'precond': PRECOND_LIST, 'step': '수정', 'category': '정산 보류 해제',
         'item': '정산 보류 해제',
         'procedure': "1. 정산 상태가 '보류(H)'인 건 체크박스 선택\n2. 하단 '보류 해제' 버튼 클릭\n3. 해제 사유 입력 팝업 노출\n4. 사유 입력 후 확인 클릭",
         'expected': "1. '보류(H)' 상태가 아닌 건 선택 시 경고 메시지\n\n2. 해제 처리 후 정산 상태 '대기(W)' 복귀\n\n3. 해제 사유 및 처리자 정보 기록\n\n4. 보류 이력은 별도 테이블에 보존",
         'log_db_type': 'update'},
        {'id': 'TB_01_005', 'precond': PRECOND_LIST, 'step': '조회', 'category': '엑셀 다운로드',
         'item': '검색 결과 엑셀 다운로드',
         'procedure': "1. 검색 조건 입력 후 조회\n2. 우측 상단 '엑셀 다운로드' 버튼 클릭\n3. 다운로드 진행 확인\n4. 다운받은 파일 열어서 데이터 확인",
         'expected': "1. 조회된 전체 건이 엑셀로 다운로드 (페이징 무관)\n\n2. 파일명: 가맹점정산내역_YYYYMMDD_HHMMSS.xlsx\n\n3. 화면 표시 컬럼과 동일 순서로 출력\n\n4. 1만 건 초과 시 경고 후 다운로드 거절",
         'log_db_type': 'select'},
        {'id': 'TB_02_001', 'precond': PRECOND_DETAIL, 'step': '조회', 'category': '정산 상세 정보 조회',
         'item': '정산 번호 클릭 시 상세 팝업',
         'procedure': "1. 목록에서 정산 번호 클릭\n2. 상세 팝업 노출 확인\n3. 정산 기본 정보, 거래 내역, 수수료 내역 탭 전환",
         'expected': "1. 정산 기본 정보 정상 노출\n - 정산번호, 가맹점, 정산일자, 정산금액 등\n\n2. 거래 내역 탭에 매핑 거래 목록 조회\n\n3. 수수료 내역 탭에 수수료 계산 내역 조회\n\n4. 정산 상태에 따라 노출 버튼 분기\n - 대기: 수정, 취소 노출\n - 완료: 조회만 가능\n - 보류: 보류 해제 노출",
         'log_db_type': 'select'},
        {'id': 'TB_02_001', 'precond': PRECOND_DETAIL, 'step': '등록', 'category': '수동 정산 등록',
         'item': '수동 정산 신규 등록',
         'procedure': "1. 목록 화면 하단 '수동 정산 등록' 버튼 클릭\n2. 등록 팝업에서 가맹점 검색 (필수)\n3. 정산 일자, 정산 금액 입력\n4. 필수 입력 누락 시 에러 메시지 확인\n5. 저장 버튼 클릭",
         'expected': "1. 정산번호 자동 채번 확인 (예: ST20250528000001)\n\n2. 정산 상태 '대기(W)' 고정\n\n3. 저장 후 목록 1행에 노출\n\n4. 등록자, 등록일시 자동 기록\n\n5. 동일 가맹점 + 동일 일자 중복 등록 시 경고",
         'log_db_type': 'insert'},
        {'id': 'TB_02_003', 'precond': PRECOND_DETAIL, 'step': '수정', 'category': '정산 정보 수정',
         'item': '대기 상태 정산 수정',
         'procedure': "1. 상세 팝업에서 '수정' 버튼 클릭\n 1-1. 정산 상태가 '대기(W)'일 때만 노출\n2. 정산 금액, 비고 등 수정\n3. 저장 버튼 클릭",
         'expected': "1. 수정 데이터 정상 저장\n\n2. 수정자, 수정일시 자동 기록\n\n3. 변경 이력 별도 테이블 저장\n - 변경 전/후 값, 변경자, 변경일시",
         'log_db_type': 'update'},
        {'id': 'TB_02_004', 'precond': PRECOND_DETAIL, 'step': '수정', 'category': '정산 취소',
         'item': '정산 취소 처리',
         'procedure': "1. 상세 팝업에서 '취소' 버튼 클릭\n 1-1. 정산 상태가 '대기(W)' 또는 '보류(H)'일 때만 노출\n2. 취소 사유 입력 팝업 노출 (필수)\n3. 사유 입력 후 확인 클릭",
         'expected': "1. 정산 상태 '취소(X)' 변경\n\n2. 목록에서 회색 표시\n\n3. 취소 사유, 취소자, 취소일시 기록\n\n4. 취소 후 수정/재취소 불가 (조회만 가능)",
         'log_db_type': 'update'},
    ]

    wb, file_name = build_workbook("PGBO-1001", "가맹점 정산 내역 관리 화면 개발 요청", sample_cases, system_type='BO')
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs')
    out_path = save_workbook(wb, file_name, out_dir)
    print(f'Saved: {out_path}')
    print(f'Size: {os.path.getsize(out_path)} bytes')
